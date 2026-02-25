"""Generate the Fast Track Tool Classification Excel Report."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tool Classification"

# ── Styles ──────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
module_fill = PatternFill(start_color="FFF469", end_color="FFF469", fill_type="solid")
module_font = Font(name="Calibri", bold=True, size=11)
body_font = Font(name="Calibri", size=10)
wrap = Alignment(wrap_text=True, vertical="top")
center_wrap = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Color-code for timing
fill_before = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # green
fill_during = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # blue
fill_both   = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # orange

# ── Column widths ───────────────────────────────────────────────
columns = [
    ("Module", 18),
    ("Tool #", 7),
    ("Tool Name", 22),
    ("Tool Type", 14),
    ("When To Complete", 16),
    ("Purpose", 50),
    ("Tool Logic\n(What the user does step-by-step)", 60),
    ("Final Decision\n(What the team decides at the end)", 45),
    ("Audit Status", 12),
]

for i, (name, width) in enumerate(columns, 1):
    col_letter = get_column_letter(i)
    ws.column_dimensions[col_letter].width = width
    cell = ws.cell(row=1, column=i, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_wrap
    cell.border = thin_border

ws.row_dimensions[1].height = 36

# ── Data ────────────────────────────────────────────────────────
tools = [
    # Module 0
    {
        "module": "Module 0: Intro Sprint",
        "num": "00",
        "name": "WOOP",
        "type": "Reflection",
        "when": "Before Meeting",
        "purpose": "Set personal commitment for the Fast Track journey using the Wish-Outcome-Obstacle-Plan framework. Establishes psychological contract.",
        "logic": "1. Define your Wish (what you want from Fast Track)\n2. Visualize the best Outcome\n3. Identify the main Obstacle\n4. Create an If-Then Plan to overcome it\n5. Rate commitment level\n6. Submit to AI coach for review",
        "decision": "Personal commitment: \"I will commit to this program because [reason], and when [obstacle] happens, I will [plan].\"",
        "status": "PASS",
    },
    # Module 1
    {
        "module": "Module 1: Identity",
        "num": "01",
        "name": "Know Thyself",
        "type": "Reflection",
        "when": "Before Meeting",
        "purpose": "Self-assess leadership style across 6 dimensions. Produces a personal leadership profile with strengths and blind spots.",
        "logic": "1. Rate yourself 1-10 on 6 leadership dimensions (delegation, decision-making, communication, time management, EQ, strategic thinking)\n2. For each: describe what you do well and what needs work\n3. Identify top 3 development priorities\n4. Review leadership profile radar chart\n5. Submit for AI review",
        "decision": "\"My top 3 leadership development priorities are [X, Y, Z] because these have the biggest impact on my team.\"",
        "status": "PASS",
    },
    {
        "module": "Module 1: Identity",
        "num": "02",
        "name": "Dream",
        "type": "Canvas / Builder",
        "when": "Before Meeting",
        "purpose": "Craft the company's 10-year vision. Goes from urgency (Corporate Killer) through dream questions to a one-page narrative, one-sentence distillation, and visual dream board.",
        "logic": "1. Corporate Killer: What would be lost if the company disappeared?\n2. Employee Promise: What employees get (money, pride, freedom)\n3. 6 Dream Questions (achievement, customers say, employees say, market position, office feels like, personal achievement)\n4. Write one-page Dream Narrative in present tense\n5. Distill into one bold sentence\n6. Dream Board: 7-10 words + 7-10 figures + 7-10 images\n7. Dream Strength Analyzer reviews quality",
        "decision": "\"Our company Dream in one sentence is: [sentence]. This is our 10-year north star for every strategic decision.\"",
        "status": "FIXED",
    },
    {
        "module": "Module 1: Identity",
        "num": "03",
        "name": "Values",
        "type": "Builder",
        "when": "Before Meeting",
        "purpose": "Define 5 core values and translate them into observable daily behaviors, red lines, recruitment questions, and a rollout plan.",
        "logic": "1. Identify 5 core company values from behaviors\n2. Gap Analysis: Compare values against Dream (auto-loaded)\n3. Cool/Not Cool behaviors for each value\n4. Red Lines (yellow = warning, red = zero tolerance)\n5. Recruitment Questions (test values in interviews)\n6. Communication & Rollout Plan",
        "decision": "\"Our 5 core values are [list]. Each value has defined Cool/Not Cool behaviors and Red Lines that we will enforce starting [date].\"",
        "status": "PASS",
    },
    {
        "module": "Module 1: Identity",
        "num": "04",
        "name": "Team",
        "type": "Assessment",
        "when": "Before Meeting",
        "purpose": "Evaluate team health using Lencioni's 5 Dysfunctions (Trust, Conflict, Commitment, Accountability, Results). Produces radar chart and improvement plan.",
        "logic": "1. Set up team members\n2. Score each member 1-10 on 5 dysfunction dimensions\n3. View Team Summary with radar chart visualization\n4. Identify weakest dimensions\n5. Create Improvement Plan per dysfunction\n6. Action Plan with owners and deadlines\n7. Commitment Lock (team signs off)\n8. Final Review + PDF export",
        "decision": "\"Our weakest team dimension is [X]. We will address it by [actions] with [owner] responsible by [date].\"",
        "status": "PASS",
    },
    {
        "module": "Module 1: Identity",
        "num": "05",
        "name": "Fit",
        "type": "Assessment",
        "when": "Both (Individual + Team)",
        "purpose": "Analyze individual and team FIT across Values, Job, Team, and Boss dimensions. Team mode includes ABC Matrix classification.",
        "logic": "Individual path:\n1. Context: Company background\n2. Values & Job Fit: Sliders 0-100 with evidence\n3. Team & Boss Fit: Sliders 0-100 with evidence\n4. Action Plan for gaps\n\nTeam path:\n1. Team Setup: Add all members\n2. Assessment per member\n3. ABC Matrix: Classify A/B/C players",
        "decision": "\"Based on FIT analysis, [person] is an [A/B/C] player. A-players: retain. B-players: develop within [months]. C-players: transition out.\"",
        "status": "PASS",
    },
    # Module 2
    {
        "module": "Module 2: Performance",
        "num": "06",
        "name": "Cash Flow",
        "type": "Calculator",
        "when": "Before Meeting",
        "purpose": "Model cash flow across 2 years. Power of One simulator shows how small changes in 7 levers dramatically impact profit and cash.",
        "logic": "1. Setup: Select business model (manufacturing/service/retail)\n2. Year 1 Data: Revenue, COGS, operating expenses, receivables, payables, inventory\n3. Year 2 Data: Same metrics\n4. Power of One Simulator: 9 interactive sliders (price +1%, volume +1%, COGS -1%, etc.)\n5. View impact on profit and cash flow\n6. Action Plan: Assign owners to each lever",
        "decision": "\"The 3 highest-impact cash levers for us are [X, Y, Z]. Each person is assigned to improve their lever by [%] within [timeframe].\"",
        "status": "PASS",
    },
    {
        "module": "Module 2: Performance",
        "num": "07",
        "name": "Energy",
        "type": "Reflection / Builder",
        "when": "Before Meeting",
        "purpose": "Audit personal energy across 4 pillars (Sleep, Nutrition, Movement, Mental Wisdom). Design habit loops using Atomic Habits framework.",
        "logic": "1. Sleep Pillar: Rate 1-10, what you do well/not well, goals, design habit loop (trigger→routine→reward)\n2. Nutrition Pillar: Same structure\n3. Movement Pillar: Same structure\n4. Brain/Focus transition\n5. Mental Wisdom: What you can/cannot control, stop/do-less/do-more, event-gap-response analysis\n6. Canvas summary",
        "decision": "\"My 3 key energy habits are: [habit 1 with trigger→routine→reward], [habit 2], [habit 3]. My accountability partner is [name].\"",
        "status": "PASS",
    },
    {
        "module": "Module 2: Performance",
        "num": "08",
        "name": "Goals",
        "type": "Brainstorm / Calculator",
        "when": "Before Meeting",
        "purpose": "Brainstorm actions and prioritize using the Impact/Easy Matrix. Applies 80/20 principle to focus on vital few high-impact activities.",
        "logic": "1. Brainstorm up to 20 potential actions/ideas\n2. Rate each on Impact (1-3) and Ease (1-3)\n3. Auto-calculate score (Impact + Ease)\n4. View ranked list sorted by score\n5. Top items = priorities for the quarter\n6. Submit for AI coach review",
        "decision": "\"Our top 5 priorities for this quarter are [ranked list]. We will focus obsessively on these before anything else.\"",
        "status": "PASS",
    },
    {
        "module": "Module 2: Performance",
        "num": "09",
        "name": "Focus",
        "type": "Planning / Builder",
        "when": "Before Meeting",
        "purpose": "Build a complete personal productivity system: brain dump, daily plan, eliminate distractions, build routines, and prime flow state.",
        "logic": "1. Empty Brain & Impact/Easy: Dump 10 tasks → categorize in 2x2 matrix\n2. Daily Plan: Top 3 priorities + time-block schedule by energy peaks\n3-4. Eliminate/Limit/Delegate (Part 1 & 2): Ruthlessly cut low-value activities\n5. Reduce Distractions: Digital + physical action plan\n6. Key Daily Routines: Morning, work start, breaks, end-of-day\n7. 3 Routines Builder: Trigger→Action→Reward habit loops\n8. Flow State Priming: Goals, challenge/skill balance, environment, pre-flow ritual\nTeam: Focus & Productivity Dynamics (WWW analysis)",
        "decision": "\"My golden hour is [time]. My 3 daily habits are [list]. I will eliminate [activities] and delegate [tasks] to [people].\"",
        "status": "PASS",
    },
    {
        "module": "Module 2: Performance",
        "num": "10",
        "name": "Performance",
        "type": "Assessment / Planning",
        "when": "Both (Individual + Team)",
        "purpose": "Build performance tracking and accountability systems. Individual: process tools (5 Whys, 80/20, Impact-Easy, Accountability). Team: dashboards and consequences.",
        "logic": "Individual path (4 steps):\n1. 5 Whys: Define problem → ask why 5 times → identify root cause → solution\n2. 80/20 Principle: Define objective → list inputs → rank by impact → focus on vital few\n3. Brainstorming & Impact-Easy: 8 ideas → 4-quadrant matrix\n4. Accountability Framework: Expectations, ownership, checkpoints, consequences\n\nTeam path (2 steps):\n1. Performance Analysis System: Binary task tracking (done/not done) + dashboard metrics\n2. Accountability System: Roles, framework, transparency, timeline",
        "decision": "\"We will track tasks as done/not done weekly. Root causes for missed tasks will be analyzed using 5 Whys. Consequences for accountability: [rewards for success, corrective for gaps].\"",
        "status": "PASS",
    },
    {
        "module": "Module 2: Performance",
        "num": "11",
        "name": "Meeting Rhythm",
        "type": "Planning",
        "when": "Before Meeting",
        "purpose": "Design the company's meeting system: priorities breakdown, meeting readiness, evaluation, and Big Rock discussion framework.",
        "logic": "1. Individual Priorities Breakdown: Break quarterly priorities into weekly tasks with hours\n2. Meeting Readiness Checklist: Purpose, participants, agenda items (topic/owner/duration), pre-materials, roles (facilitator/scribe/timekeeper)\n3. Meeting Evaluation: Rate effectiveness, what worked/didn't, Start/Stop/Do Better\n4. Big Rock Discussion: Problem statement → root cause → solutions → WWW (Who/What/When)",
        "decision": "\"Our meeting rhythm is: [daily/weekly/quarterly schedule]. Each meeting has a clear purpose, max 8 participants, and ends with WWW commitments.\"",
        "status": "PASS (minor)",
    },
    # Module 3
    {
        "module": "Module 3: Strategy",
        "num": "12",
        "name": "Market Size",
        "type": "Calculator",
        "when": "Before Meeting",
        "purpose": "Calculate Total Addressable Market and Profit Pool. Analyze driving forces that will change market size in the future.",
        "logic": "1. Market Size Inputs: Total customers × Average spending × Purchases per year\n2. Auto-calculate Market Size\n3. Gross Margin input → Auto-calculate Profit Pool\n4. Brainstorm Driving Forces (trends, regulations, technology, etc.)\n5. Rate each force: Impact (1-5) and Probability (1-5)\n6. Prioritize top forces (score = 2×Impact + Probability)\n7. Future Impact: Model how top forces change customers/spending/purchases/margin",
        "decision": "\"Our TAM is [€X]. Our addressable Profit Pool is [€Y]. The top 3 forces that will change this are [list]. In [N] years, market could be [€Z].\"",
        "status": "PASS",
    },
    {
        "module": "Module 3: Strategy",
        "num": "13",
        "name": "Segmentation",
        "type": "Assessment / Calculator",
        "when": "Before Meeting",
        "purpose": "Define market segments behaviorally, quantify them, and rank using a weighted Attractiveness Index to choose where to compete.",
        "logic": "1. Define Segments: Describe customer behaviors → create 3+ segments (name, description, behaviors)\n2. Quantify Segments: Total market, accessible market, avg revenue, total value per segment\n3. Attractiveness Index: Rate each segment on Size (25%), Growth (25%), Profitability (30%), Fit (20%) → auto-calculate weighted score\n4. Target Market & Goals: Select top 2-3 segments, set market share goals with timeline and strategy, define end game vision",
        "decision": "\"We will target [2-3 segments] ranked by attractiveness. Our market share goal for [segment 1] is [X%] by [date]. Our end game is [vision].\"",
        "status": "PASS",
    },
    {
        "module": "Module 3: Strategy",
        "num": "14",
        "name": "Target Segment",
        "type": "Validation / Builder",
        "when": "Before Meeting",
        "purpose": "Deep-dive into target segment through customer interviews. Validate pains, needs, and gains using prioritization matrices.",
        "logic": "1. Interview Guide: Define target segment, objective, create 3+ questions (Mom Test principles)\n2. Interview Notes: Capture 1-5 interviews (customer name, profile, pains, needs, gains, quotes, observations)\n3. Prioritization Matrices: Score pains/needs/gains on Severity × Frequency × Impact\n4. Summary: Core customer profile, top 3 pains/needs/gains, key insights, strategic implications",
        "decision": "\"Our core customer is [profile]. Their #1 pain is [X], #1 need is [Y], #1 desired gain is [Z]. This shapes our value proposition.\"",
        "status": "PASS",
    },
    # Module 4
    {
        "module": "Module 4: Strategy Dev",
        "num": "15",
        "name": "Value Proposition",
        "type": "Builder / Validation",
        "when": "Before Meeting",
        "purpose": "Build the VP using the Fast Track formula and validate with the 8-point checklist. Creates the core strategic statement for all downstream tools.",
        "logic": "1. Pain Map: Target customer, biggest pain, evidence, current solutions, breaking point, desired outcome\n2. Your Edge: Map 1-5 competitors (strength/weakness/differentiation), define unfair advantage, why they'd switch\n3. The Statement: Build VP formula live: \"For [who] who [pain], we provide [solution] that [benefit], unlike [competitor gap]\"\n4. Anti-promise: What we refuse to promise\n5. Stress Test: 8-point validation checklist (6/8 required to pass)\n6. 48-hour deployment plan + refinement plan",
        "decision": "\"Our VP is: [one sentence]. It scores [X/8] on the validation checklist. We will test it with 10 target buyers within 48 hours.\"",
        "status": "PASS",
    },
    {
        "module": "Module 4: Strategy Dev",
        "num": "16",
        "name": "VP Testing",
        "type": "Validation",
        "when": "Both (Individual + Team)",
        "purpose": "Two-round customer validation of the VP. Round 1 discovers reactions, Round 2 validates the refined message. Evidence-based decisions only.",
        "logic": "1. Test Setup: VP being tested, objective, target customer profile, interview plan\n2. Round 1 Interviews: Capture reactions (problem reconfirm, value reaction, what worked/unclear, concerns, their words, call-to-action response)\n3. Analyze & Refine: Identify patterns (positive, confusion, objections) → refine VP with justification\n4. Round 2 Validation: Score clarity, enthusiasm, differentiation, actionability (1-5)\n5. Final Decision: Lock VP + 7-item sprint checklist + rollout plan",
        "decision": "\"After [N] customer interviews, our validated VP is: [refined sentence]. Pattern: [X]% understood it immediately. We go/no-go on [date].\"",
        "status": "PASS",
    },
    {
        "module": "Module 4: Strategy Dev",
        "num": "17",
        "name": "Product Development",
        "type": "Assessment / Builder",
        "when": "Both (Individual + Team)",
        "purpose": "Build the product portfolio using WTP (Willingness to Pay) customer testing. Categorize into Storytellers, Top Line, and Profit Contributors.",
        "logic": "1. Feature Identification: List 10-15 features (name, description, current state)\n2. Customer Interview Setup: Target customers, interview count, format, key questions\n3. WTP Testing Results: Customer value feedback, willingness to pay, competitive advantage, priority score (1-10)\n4. Product Portfolio Design: Categorize products as Storytellers (<10%), Top Line (70-80% sales), Profit (high margin)\n5. Implementation Plan: Quick wins (30 days), resources, timeline, success metrics",
        "decision": "\"Our portfolio: [N] Storytellers (brand builders), [N] Top Line (volume drivers), [N] Profit (margin protectors). We kill/keep/invest in [specific products].\"",
        "status": "PASS",
    },
    {
        "module": "Module 4: Strategy Dev",
        "num": "18",
        "name": "Pricing",
        "type": "Builder / Calculator",
        "when": "Before Meeting",
        "purpose": "Design pricing strategy: classify features, set price anchor, create tier bundles, align price-to-value, and plan pilot.",
        "logic": "1. Feature Classification: Categorize 5+ features as Must-Have, Nice-to-Have, or Killer with WTP impact\n2. Price Anchor: Choose anchor position (premium/mid/entry), define target segment, competitor benchmark, price rationale\n3. Tier Bundle Design: Create 2-3 tiers (name, price, features, target customer, margin)\n4. Price-to-Value Alignment: Map must-have/nice-to-have/killer features to tiers\n5. Implementation: 3+ pricing actions with owners, pilot plan, success metrics",
        "decision": "\"Our anchor is [premium/mid/entry] at [€X]. We offer [N] tiers: [Basic at €A, Pro at €B, Premium at €C]. Pilot starts [date] with [segment].\"",
        "status": "PASS",
    },
    # Module 5
    {
        "module": "Module 5: Go to Market",
        "num": "19",
        "name": "Brand Marketing",
        "type": "Builder",
        "when": "Before Meeting",
        "purpose": "Build brand from inside out: promise, personality, cult brand model (employee + consumer), messaging pillars, and campaign plan.",
        "logic": "1. Brand Promise: One-sentence promise + emotional connection + VP transformation + brand enemy (what you're NOT)\n2. Brand Personality: 3-5 traits with descriptions and examples + brand voice + brand archetype\n3. Cult Brand Model: Employee brand (values, story) + Consumer brand (experience, shared values)\n4. Marketing Strategy: 3+ messaging pillars with rationale and channels + content strategy\n5. Implementation: 3+ marketing actions + success metrics + consistency checklist",
        "decision": "\"Our brand promise is: [sentence]. Our brand enemy is [X]. Our 3 messaging pillars are [list]. Campaign launches [date] across [channels].\"",
        "status": "PASS",
    },
    {
        "module": "Module 5: Go to Market",
        "num": "20",
        "name": "Customer Service",
        "type": "Canvas / Planning",
        "when": "Before Meeting",
        "purpose": "Map customer journey, design WOW moments and Just Better improvements, assign ownership with KPIs, and create implementation plan.",
        "logic": "1. Journey Map: Define 5-8 stages (awareness→exit) with customer goal, action, and thought per stage\n2. WOW Touchpoints: 3 memorable, emotional, share-worthy moments + measurement\n3. Just Better Improvements: 3 improvements in speed/clarity/tone with execution plan\n4. Ownership & Metrics: One owner + KPI + deadline per touchpoint\n5. Implementation: Execution plan, frontline training, metric tracking, quarterly review",
        "decision": "\"Our 3 WOW moments are [list]. Our 3 Just Better improvements are [list]. Each has one owner and one KPI. First review in [weeks].\"",
        "status": "PASS",
    },
    {
        "module": "Module 5: Go to Market",
        "num": "21",
        "name": "Route to Market",
        "type": "Assessment / Planning",
        "when": "Before Meeting",
        "purpose": "Choose the 1-2 most effective routes to market by filtering options through VP-RTM fit, scenario testing, and execution planning.",
        "logic": "1. RTM Options Analysis: List 3-5 options rated on reach, margin, control, brand, operational load\n2. VP-RTM Fit Filter: 5 yes/no questions (delivers promise? reaches customer? protects margin? scales? fits ops?) → auto-score\n3. Scenario Testing: Model top 2-3 options (capabilities, people, tools, investment, timeline, risks)\n4. Final Decision: Select primary + secondary RTM with justification and success metrics\n5. Execution Plan: 30-day actions, resources, ownership, checkpoint, deal-breakers",
        "decision": "\"Our primary RTM is [X] (score [N/5]). Secondary: [Y]. We rejected [Z] because [reason]. First checkpoint: [date].\"",
        "status": "PASS",
    },
    # Module 6
    {
        "module": "Module 6: Execution",
        "num": "22",
        "name": "Core Activities",
        "type": "Brainstorm / Assessment",
        "when": "Before Meeting",
        "purpose": "Identify the top 5 core activities that drive competitive advantage. Define 3 processes per activity with KPIs.",
        "logic": "1. Value Proposition Input: Start from your VP statement\n2. Brainstorm Activities: List 5-10+ potential activities (start with a verb)\n3. Select Top 5: Apply 80/20 principle to choose the 5 most impactful\n4. Confirm Selection: Review and finalize top 5\n5. Processes & KPIs: Define 3 core processes per activity + KPI (name, unit, target) for each",
        "decision": "\"Our 5 core activities are [list]. Each has 3 defined processes and measurable KPIs. This is our execution engine.\"",
        "status": "PASS",
    },
    {
        "module": "Module 6: Execution",
        "num": "23",
        "name": "Processes & Decisions",
        "type": "Planning / Assessment",
        "when": "Before Meeting",
        "purpose": "Map core decisions, data needs, KPIs, and capabilities for each activity. Identify gaps and create development plans.",
        "logic": "1. Core Decisions: For each activity, define 3-5 key decisions (decision, frequency, impact level)\n2. Decision Data & KPIs: What data is needed + 2 KPIs per activity (metric, target)\n3. Required Capabilities: Human (skills), Systemic (tools/workflows), Relational (partnerships)\n4. Gap Analysis: Current capabilities vs. required → identify gaps → development plan with timelines",
        "decision": "\"For each core activity, we have defined the key decisions, required data, KPIs, and capability gaps. Development plan addresses [top gaps] by [quarter].\"",
        "status": "PASS",
    },
    {
        "module": "Module 6: Execution",
        "num": "24",
        "name": "FIT ABC",
        "type": "Assessment / Planning",
        "when": "Before Meeting",
        "purpose": "Rate each team member on 4 FIT dimensions, classify as A/B/C player, and create tailored action plans (retain/develop/transition).",
        "logic": "1. Team Members: Create evaluation roster (name, role, tenure)\n2. FIT Analysis: Rate each person on Energy, Skill, Values, Impact (1-5 scale) with notes\n3. ABC Classification: A (great + cultural fit + potential), B (good but not great), C (low/misfit) with 50-char reasoning\n4. Action Plans: A-players (coaching, role growth, timeline), B/C-players (development needs, tough decisions)",
        "decision": "\"Team is [N]% A-players, [N]% B-players, [N]% C-players. A-players: [retain plan]. B-players: [develop by date]. C-players: [transition plan].\"",
        "status": "PASS",
    },
    # Module 7
    {
        "module": "Module 7: Org Design",
        "num": "25",
        "name": "Org Redesign",
        "type": "Planning",
        "when": "During Meeting",
        "purpose": "Redesign organizational structure: map current team with FIT scores, define new roles, match people to roles via decision matrix, create PDPs.",
        "logic": "1. Current Team Assessment: Map people with 4-D FIT scores (Company, Job, Team, Boss)\n2. Define New Roles: Title, criticality (1-10), 3+ required skills per role\n3. Decision Matrix: Score each person against each role for best fit\n4. Action Plan: Auto-categorize into immediate assignments, development, transitions\n5. Personal Development Plans: Target role, core activity, capabilities with completion methods and timelines",
        "decision": "\"New org structure has [N] roles. [People] move immediately, [people] need [months] development, [people] transition out. Each has a PDP.\"",
        "status": "PASS (minor)",
    },
    {
        "module": "Module 7: Org Design",
        "num": "26",
        "name": "Employer Branding",
        "type": "Builder",
        "when": "Before Meeting",
        "purpose": "Build Employer Value Proposition and recruitment playbook: segment talent, rank needs, formulate EVP, test authenticity, design campaign.",
        "logic": "1. Employee Segmentation: Target segment, roles, mindset, culture preferences\n2. Needs Ranking: Rank 5 categories (Compensation, Benefits, Career, Work Environment, Culture) 1-5\n3. EVP Formulation: Primary need, core promise, differentiator, EVP statement\n4. EVP Testing: 3+ internal truths with evidence + 2+ anti-EVP statements + credibility check\n5. Campaign & Recruitment: Brand message, tagline, channels (inbound/outbound/content), recruitment process (CV→Phone→Assessment→Interview→Scorecard), campaign activation",
        "decision": "\"Our EVP is: [statement]. Anti-EVP: We are NOT [list]. Recruitment process: 5-step scorecard. Campaign launches [date] on [channels].\"",
        "status": "PASS",
    },
    {
        "module": "Module 7: Org Design",
        "num": "27",
        "name": "Agile Teams",
        "type": "Brainstorm / Planning",
        "when": "Before Meeting",
        "purpose": "Form agile teams to tackle top problems/opportunities. Brainstorm issues, prioritize with Impact/Ease, create team charters.",
        "logic": "1. Brainstorm: List 8+ problems and opportunities (description + type)\n2. Prioritize: Rate each on Impact (1-3) and Ease (1-3), auto-rank by score\n3. Charter Creation: For top 3 items, create team charters with: project goal, specific metric, timeline, 2+ team roles with responsibilities, 2+ KPIs with targets and tracking frequency\n4. Review & Submit: Summary of 3 charters + next steps checklist (sprint-zero meetings, review cadence, dashboard, calendar)",
        "decision": "\"We launch 3 agile teams: [Team 1] tackling [problem] by [date], [Team 2] for [opportunity], [Team 3] for [issue]. Sprint-zero meetings start [date].\"",
        "status": "PASS",
    },
    # Module 8
    {
        "module": "Module 8: Technology",
        "num": "28",
        "name": "Digitalization",
        "type": "Assessment / Planning",
        "when": "Both (Individual + Team)",
        "purpose": "Identify top 3 decisions and top 3 processes for AI/digitalization. Individual audit feeds into team strategy and 3-week implementation plan.",
        "logic": "Individual path:\n1. Decision Inventory: List decisions with frequency, impact, data needs, AI opportunities\n2. Process Inventory: List processes with time consumed, repetition, automation potential, KPIs\n3. AI Implementation Planning: Tools, costs, timeline, success metrics, obstacles\n\nTeam path:\n1. Collect individual analyses\n2. Pattern Recognition: Common decisions, processes, tools, collaboration opportunities\n3. Strategy Development: Connect to company goals, prioritize, cross-functional opportunities\n4. Implementation: Week 1-3 actions with owners and deadlines",
        "decision": "\"Our top 3 decisions for AI are [list]. Top 3 processes for automation are [list]. Week 1 actions: [quick wins]. Full rollout by [date].\"",
        "status": "PASS",
    },
    {
        "module": "Module 8: Technology",
        "num": "29",
        "name": "Digital Heart",
        "type": "Builder / Planning",
        "when": "Before Meeting",
        "purpose": "Design the company's data lake architecture: core decisions, data sources, technology platform, APIs, and Phase 1 implementation.",
        "logic": "1. Core Decisions: Identify decisions for data lake (name, frequency, impact, current process, data needed, priority 1-10)\n2. Data Sources: Map internal + external sources per decision, assess quality and gaps\n3. Technology Platform: Collection method, storage system, algorithms, processing frequency, data integration, security\n4. APIs & User Access: Endpoints, access method, user interface, update frequency, training, access controls\n5. Implementation: Phase 1 timeline, actions, quick wins, resources, estimated cost, success metrics",
        "decision": "\"Our Digital Heart connects [N] data sources into one platform. Phase 1 delivers [quick wins] by [date]. Investment: [€X]. Success metric: decision time from [weeks] to [hours].\"",
        "status": "PASS",
    },
]

# ── Write data rows ─────────────────────────────────────────────
current_module = ""
row = 2
for t in tools:
    # Module header row
    if t["module"] != current_module:
        current_module = t["module"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
        cell = ws.cell(row=row, column=1, value=current_module)
        cell.font = module_font
        cell.fill = module_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws.row_dimensions[row].height = 24
        row += 1

    values = [
        "",  # module col empty for data rows
        t["num"],
        t["name"],
        t["type"],
        t["when"],
        t["purpose"],
        t["logic"],
        t["decision"],
        t["status"],
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = thin_border

    # Center the tool number and status
    ws.cell(row=row, column=2).alignment = center_wrap
    ws.cell(row=row, column=9).alignment = center_wrap

    # Color-code timing column
    timing_cell = ws.cell(row=row, column=5)
    timing_cell.alignment = center_wrap
    if "Both" in t["when"]:
        timing_cell.fill = fill_both
    elif "During" in t["when"]:
        timing_cell.fill = fill_during
    else:
        timing_cell.fill = fill_before

    # Color-code status
    status_cell = ws.cell(row=row, column=9)
    if "PASS" in t["status"] and "minor" not in t["status"] and "FIXED" not in t["status"]:
        status_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    elif "FIXED" in t["status"]:
        status_cell.fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    elif "minor" in t["status"]:
        status_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    # Set row height based on content
    ws.row_dimensions[row].height = max(80, len(t["logic"].split("\n")) * 14)
    row += 1

# ── Freeze header row ───────────────────────────────────────────
ws.freeze_panes = "A2"

# ── Auto-filter ─────────────────────────────────────────────────
ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{row - 1}"

# ── Save ────────────────────────────────────────────────────────
output_path = r"C:\Users\Admin\Desktop\show time\docs\audits\Fast-Track-Tool-Classification.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
